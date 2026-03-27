#!/usr/bin/env python3
import argparse
import json
from pathlib import Path
from openpyxl import load_workbook, Workbook


def upsert_manifest(manifest_path: Path, entry: dict):
    rows = []
    if manifest_path.exists():
        for ln in manifest_path.read_text(encoding='utf-8').splitlines():
            if ln.strip():
                rows.append(json.loads(ln))

    updated = False
    for i, r in enumerate(rows):
        if str(r.get('article_id', '')) == str(entry['article_id']):
            rows[i] = entry
            updated = True
            break
    if not updated:
        rows.append(entry)

    manifest_path.parent.mkdir(parents=True, exist_ok=True)
    manifest_path.write_text(''.join(json.dumps(r, ensure_ascii=False) + '\n' for r in rows), encoding='utf-8')


def upsert_index(xlsx_path: Path, entry: dict):
    headers = ['article_date', 'provider', 'article_id', 'article_title', 'crawl_date', '类型', '行业', '标签', '作者', '联合发布人']

    if xlsx_path.exists():
        wb = load_workbook(xlsx_path)
        ws = wb[wb.sheetnames[0]]
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = '索引'
        ws.append(headers)

    # ensure header
    existing_headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    if ws.max_row == 0:
        ws.append(headers)
    else:
        # non-destructive header migration: append missing columns instead of wiping rows
        for h in headers:
            if h not in existing_headers:
                ws.cell(1, ws.max_column + 1).value = h
                existing_headers.append(h)

    col = {}
    for c in range(1, ws.max_column + 1):
        h = ws.cell(1, c).value
        if h:
            col[str(h)] = c
    for h in headers:
        if h not in col:
            ws.cell(1, ws.max_column + 1).value = h
            col[h] = ws.max_column

    target_row = None
    last_filled_row = 1
    for r in range(2, ws.max_row + 1):
        v = ws.cell(r, col['article_id']).value
        if v not in (None, ''):
            last_filled_row = r
        if str(v) == str(entry['article_id']):
            target_row = r
            break

    if target_row is None:
        target_row = last_filled_row + 1

    ws.cell(target_row, col['article_date']).value = entry.get('article_date', '')
    ws.cell(target_row, col['provider']).value = entry.get('provider', '')
    ws.cell(target_row, col['article_id']).value = entry.get('article_id', '')
    ws.cell(target_row, col['article_title']).value = entry.get('article_title', '')
    ws.cell(target_row, col['crawl_date']).value = entry.get('crawl_date', '')
    ws.cell(target_row, col['类型']).value = entry.get('content_type', '')
    ws.cell(target_row, col['行业']).value = entry.get('industry', '')
    ws.cell(target_row, col['标签']).value = ', '.join(entry.get('tags', []))
    ws.cell(target_row, col['作者']).value = entry.get('author', '')
    ws.cell(target_row, col['联合发布人']).value = entry.get('co_publisher', '')

    xlsx_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(xlsx_path)


def upsert_all(entry: dict):
    ws_root = Path(__file__).resolve().parents[3]
    manifest = ws_root / 'acecamp-raw' / 'index' / 'manifest.jsonl'
    index_xlsx = ws_root / 'acecamp-raw' / 'index' / '索引.xlsx'
    upsert_manifest(manifest, entry)
    upsert_index(index_xlsx, entry)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--article-id', required=True)
    ap.add_argument('--article-date', required=True)
    ap.add_argument('--provider', default='acecamptech')
    ap.add_argument('--article-title', required=True)
    ap.add_argument('--crawl-date', required=True)
    ap.add_argument('--source-url', required=True)
    ap.add_argument('--record-time', required=True)
    ap.add_argument('--source-path', required=True)
    ap.add_argument('--rendered-path', required=True)
    ap.add_argument('--content-type', default='')
    ap.add_argument('--industry', default='')
    ap.add_argument('--tags', default='')
    ap.add_argument('--author', default='')
    ap.add_argument('--co-publisher', default='')
    args = ap.parse_args()

    tags = [t.strip() for t in args.tags.replace('，', ',').split(',') if t.strip()]

    entry = {
        'article_date': args.article_date,
        'provider': args.provider,
        'article_id': str(args.article_id),
        'article_title': args.article_title,
        'crawl_date': args.crawl_date,
        'source_url': args.source_url,
        'record_time': args.record_time,
        'source_path': args.source_path,
        'rendered_path': args.rendered_path,
        'status': 'ok',
        'content_type': args.content_type,
        'industry': args.industry,
        'tags': tags,
        'author': args.author,
        'co_publisher': args.co_publisher,
    }

    upsert_all(entry)
    print('synced', args.article_id)


if __name__ == '__main__':
    main()
